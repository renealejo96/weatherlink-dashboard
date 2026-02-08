from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
import os
import time
import pytz
from dotenv import load_dotenv
from weatherlink_client import WeatherLinkClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# Deshabilitar caché para HTML
@app.after_request
def add_header(response):
    if response.mimetype == 'text/html':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Importar funciones de Supabase
try:
    from supabase_api import (
        get_latest_readings, 
        get_station_history, 
        get_daily_summary,
        get_all_stations_comparison
    )
    SUPABASE_ENABLED = True
except ImportError:
    SUPABASE_ENABLED = False
    print("⚠️  Supabase no disponible - usando solo API directa de WeatherLink")

# Configurar las 3 estaciones
STATIONS = {
    'finca1': {
        'name': 'PYGANFLOR',
        'api_key': os.getenv('FINCA1_API_KEY'),
        'api_secret': os.getenv('FINCA1_API_SECRET'),
        'station_id': os.getenv('FINCA1_STATION_ID')
    },
    'finca2': {
        'name': 'Urcuquí',
        'api_key': os.getenv('FINCA2_API_KEY'),
        'api_secret': os.getenv('FINCA2_API_SECRET'),
        'station_id': os.getenv('FINCA2_STATION_ID')
    },
    'finca3': {
        'name': 'Malchinguí',
        'api_key': os.getenv('FINCA3_API_KEY'),
        'api_secret': os.getenv('FINCA3_API_SECRET'),
        'station_id': os.getenv('FINCA3_STATION_ID')
    }
}

# Crear clientes para cada estación
clients = {}
for key, station in STATIONS.items():
    clients[key] = WeatherLinkClient(
        station['api_key'],
        station['api_secret'],
        station['station_id']
    )

# Sistema de caché simple (TTL: 5 minutos)
CACHE = {}
CACHE_TTL = 300  # segundos

def get_cache_key(station_key, start_ts, end_ts):
    """Generar clave de caché"""
    return f"{station_key}_{start_ts}_{end_ts}"

def get_cached_data(cache_key):
    """Obtener datos del caché si aún son válidos"""
    if cache_key in CACHE:
        cached_time, cached_data = CACHE[cache_key]
        if time.time() - cached_time < CACHE_TTL:
            return cached_data
    return None

def set_cached_data(cache_key, data):
    """Guardar datos en el caché"""
    CACHE[cache_key] = (time.time(), data)


@app.route('/')
def index():
    """Página principal con dashboard de las 3 estaciones"""
    return render_template('index.html', stations=STATIONS, supabase_enabled=SUPABASE_ENABLED)


@app.route('/rain/events')
def rain_events_dashboard():
    """Dashboard dedicado a eventos de lluvia"""
    return render_template('rain_events.html', stations=STATIONS)


@app.route('/api/current/<station_key>')
def get_current_data(station_key):
    """Obtener datos actuales de una estación"""
    if station_key not in clients:
        return jsonify({'error': 'Estación no encontrada'}), 404
    
    try:
        data = clients[station_key].get_current_conditions()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/historical/<station_key>')
def get_historical_data(station_key):
    """Obtener datos históricos de una estación"""
    if station_key not in clients:
        return jsonify({'error': 'Estación no encontrada'}), 404
    
    # Obtener parámetros de fecha
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    days = request.args.get('days', type=int, default=7)
    
    try:
        if start_date and end_date:
            # Usar fechas específicas
            start_timestamp = int(datetime.strptime(start_date, '%Y-%m-%d').timestamp())
            end_timestamp = int(datetime.strptime(end_date, '%Y-%m-%d').timestamp()) + 86399  # Final del día
        else:
            # Usar últimos N días
            end_timestamp = int(datetime.now().timestamp())
            start_timestamp = int((datetime.now() - timedelta(days=days)).timestamp())
        
        # Verificar caché
        cache_key = get_cache_key(station_key, start_timestamp, end_timestamp)
        cached_data = get_cached_data(cache_key)
        
        if cached_data:
            cached_data['from_cache'] = True
            return jsonify(cached_data)
        
        # Si no hay en caché, obtener de la API
        data = clients[station_key].get_historic_data(start_timestamp, end_timestamp)
        data['from_cache'] = False
        
        # Guardar en caché
        set_cached_data(cache_key, data)
        
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/station/<station_key>')
def station_detail(station_key):
    """Página de detalle de una estación con gráficos"""
    if station_key not in STATIONS:
        return "Estación no encontrada", 404
    
    return render_template('station_detail.html', 
                         station_key=station_key, 
                         station=STATIONS[station_key])


@app.route('/compare')
def compare_stations():
    """Página para comparar las 3 estaciones"""
    return render_template('compare.html', stations=STATIONS)


@app.route('/api/compare')
def get_compare_data():
    """Obtener datos de todas las estaciones para comparar"""
    days = request.args.get('days', type=int, default=7)
    
    end_timestamp = int(datetime.now().timestamp())
    start_timestamp = int((datetime.now() - timedelta(days=days)).timestamp())
    
    result = {}
    for key, client in clients.items():
        try:
            data = client.get_historic_data(start_timestamp, end_timestamp)
            result[key] = {
                'name': STATIONS[key]['name'],
                'data': data
            }
        except Exception as e:
            result[key] = {
                'name': STATIONS[key]['name'],
                'error': str(e)
            }
    
    return jsonify(result)


@app.route('/api/export/<station_key>')
def export_to_excel(station_key):
    """Exportar datos históricos a Excel"""
    if station_key not in clients:
        return jsonify({'error': 'Estación no encontrada'}), 404
    
    # Obtener parámetros de fecha
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    days = request.args.get('days', type=int, default=7)
    
    try:
        if start_date and end_date:
            start_timestamp = int(datetime.strptime(start_date, '%Y-%m-%d').timestamp())
            end_timestamp = int(datetime.strptime(end_date, '%Y-%m-%d').timestamp()) + 86399
        else:
            end_timestamp = int(datetime.now().timestamp())
            start_timestamp = int((datetime.now() - timedelta(days=days)).timestamp())
        
        # Obtener datos
        data = clients[station_key].get_historic_data(start_timestamp, end_timestamp)
        records = data.get('records', [])
        
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Meteorológicos"
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Encabezados
        headers = ['Fecha y Hora', 'Temperatura (°C)', 'Humedad (%)', 'Viento (km/h)', 
                   'Lluvia (mm)', 'Radiación Solar (W/m²)', 'DPV (kPa)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        # Ecuador timezone UTC-5
        for row_idx, record in enumerate(records, 2):
            # Fecha - Convertir timestamp Unix a hora de Ecuador
            # Los timestamps son UTC, restar 5 horas para Ecuador
            timestamp = record['timestamp']
            ecuador_dt = datetime.utcfromtimestamp(timestamp) - timedelta(hours=5)
            
            ws.cell(row=row_idx, column=1, value=ecuador_dt.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Temperatura (F -> C)
            temp_c = ((record.get('temperature', 0) - 32) * 5 / 9) if record.get('temperature') else None
            ws.cell(row=row_idx, column=2, value=round(temp_c, 2) if temp_c else '')
            
            # Humedad
            ws.cell(row=row_idx, column=3, value=record.get('humidity', ''))
            
            # Viento (mph -> km/h)
            wind_kmh = (record.get('wind_speed', 0) * 1.60934) if record.get('wind_speed') else None
            ws.cell(row=row_idx, column=4, value=round(wind_kmh, 2) if wind_kmh else '')
            
            # Lluvia en mm (el backend ya entrega mm)
            rain_mm = record.get('rain')
            ws.cell(row=row_idx, column=5, value=round(rain_mm, 2) if rain_mm else '')
            
            # Radiación Solar
            ws.cell(row=row_idx, column=6, value=record.get('solar_radiation', ''))
            
            # DPV (kPa)
            if record.get('temperature') and record.get('humidity'):
                tc = (record['temperature'] - 32) * 5 / 9
                vpsat = 0.6108 * (2.718281828 ** ((17.27 * tc) / (tc + 237.3)))
                vpactual = (record['humidity'] / 100) * vpsat
                dpv = vpsat - vpactual
                ws.cell(row=row_idx, column=7, value=round(dpv, 3))
            else:
                ws.cell(row=row_idx, column=7, value='')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        filename = f"{STATIONS[station_key]['name']}_datos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# RUTAS DE SUPABASE (datos en tiempo real)
# ============================================

@app.route('/api/supabase/latest')
def api_supabase_latest():
    """GET /api/supabase/latest - Últimas lecturas de todas las estaciones desde Supabase"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    result = get_latest_readings()
    if result['success']:
        return jsonify(result['data'])
    return jsonify({'error': result['error']}), 500


@app.route('/api/supabase/station/<station_key>/history')
def api_supabase_history(station_key):
    """GET /api/supabase/station/<key>/history?hours=24 - Historial desde Supabase"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    hours = int(request.args.get('hours', 24))
    result = get_station_history(station_key, hours)
    if result['success']:
        return jsonify(result['data'])
    return jsonify({'error': result['error']}), 500


@app.route('/api/supabase/station/<station_key>/daily')
def api_supabase_daily(station_key):
    """GET /api/supabase/station/<key>/daily?days=7 - Resumen diario desde Supabase"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    days = int(request.args.get('days', 7))
    result = get_daily_summary(station_key, days)
    if result['success']:
        return jsonify(result['data'])
    return jsonify({'error': result['error']}), 500


@app.route('/api/supabase/comparison')
def api_supabase_comparison():
    """GET /api/supabase/comparison - Comparación de todas las estaciones desde Supabase"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    result = get_all_stations_comparison()
    if result['success']:
        return jsonify(result['data'])
    return jsonify({'error': result['error']}), 500


@app.route('/api/rain/events/active')
def api_rain_events_active():
    """GET /api/rain/events/active - Eventos de lluvia activos"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    try:
        import requests
        url = f"{os.getenv('SUPABASE_URL')}/rest/v1/active_rain_events"
        headers = {
            "apikey": os.getenv('SUPABASE_KEY'),
            "Authorization": f"Bearer {os.getenv('SUPABASE_KEY')}",
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return jsonify({'success': True, 'data': response.json()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rain/events/history')
def api_rain_events_history():
    """GET /api/rain/events/history?station_key=finca1&limit=10 - Historial de eventos"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    station_key = request.args.get('station_key')
    limit = request.args.get('limit', 10, type=int)
    
    try:
        import requests
        url = f"{os.getenv('SUPABASE_URL')}/rest/v1/rain_events"
        headers = {
            "apikey": os.getenv('SUPABASE_KEY'),
            "Authorization": f"Bearer {os.getenv('SUPABASE_KEY')}",
        }
        params = {
            'order': 'event_start.desc',
            'limit': limit
        }
        if station_key:
            params['station_key'] = f'eq.{station_key}'
        
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        return jsonify({'success': True, 'data': response.json()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rain/accumulated')
def api_rain_accumulated():
    """GET /api/rain/accumulated - Lluvia acumulada por día y semana"""
    if not SUPABASE_ENABLED:
        return jsonify({'error': 'Supabase no configurado'}), 503
    
    try:
        import requests
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # Obtener todos los eventos de lluvia de las últimas 8 semanas
        url = f"{os.getenv('SUPABASE_URL')}/rest/v1/rain_events"
        headers = {
            "apikey": os.getenv('SUPABASE_KEY'),
            "Authorization": f"Bearer {os.getenv('SUPABASE_KEY')}",
        }
        
        # Fecha de hace 8 semanas
        eight_weeks_ago = (datetime.now() - timedelta(weeks=8)).isoformat()
        
        params = {
            'event_start': f'gte.{eight_weeks_ago}',
            'order': 'event_start.desc',
            'limit': 1000
        }
        
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        events = response.json()
        
        # Agrupar por estación, semana y día
        by_week = defaultdict(lambda: defaultdict(float))  # {station: {week: total}}
        by_day = defaultdict(lambda: defaultdict(float))   # {station: {date: total}}
        
        for event in events:
            station_key = event['station_key']
            station_name = event['station_name']
            rain = event.get('rain_accumulated', 0) or 0
            
            # Parsear fecha de inicio
            event_start = datetime.fromisoformat(event['event_start'].replace('Z', '+00:00'))
            
            # Calcular número de semana (formato: YY-WW)
            year = event_start.year % 100  # Últimos 2 dígitos del año
            week = event_start.isocalendar()[1]  # Número de semana
            week_key = f"{year:02d}-{week:02d}"
            
            # Fecha del día (formato: YYYY-MM-DD)
            day_key = event_start.strftime('%Y-%m-%d')
            
            # Acumular
            by_week[station_key][week_key] += rain
            by_day[station_key][day_key] += rain
        
        # Convertir a formato de respuesta
        result = {
            'by_week': {},
            'by_day': {},
            'stations': {}
        }
        
        # Agregar nombres de estaciones
        for station_key in STATIONS:
            result['stations'][station_key] = STATIONS[station_key]['name']
        
        # Convertir defaultdict a dict normal
        for station_key in by_week:
            result['by_week'][station_key] = dict(by_week[station_key])
        
        for station_key in by_day:
            result['by_day'][station_key] = dict(by_day[station_key])
        
        return jsonify({'success': True, 'data': result})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dashboard')
def dashboard():
    """Dashboard en tiempo real con datos de Supabase"""
    if not SUPABASE_ENABLED:
        return render_template('error.html', 
                             message='Supabase no está configurado. Usa la vista principal.'), 503
    
    return render_template('dashboard.html', stations=STATIONS)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
